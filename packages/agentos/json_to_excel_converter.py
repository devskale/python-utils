#!/usr/bin/env python3
"""
JSON to Excel/CSV Converter for Document Intelligence System

Converts matched document JSON files into structured Excel/CSV format
for better readability and analysis.
"""

import json
import pandas as pd
import argparse
import sys
import csv
from pathlib import Path
import credgoo
from credgoo import get_api_key
from ofs.api import list_bidder_docs_json

def convert_json_to_table(json_file_path, output_format='xlsx'):
    """
    Convert JSON document matching results to Excel/CSV table format.
    
    Args:
        json_file_path (str): Path to the JSON file
        output_format (str): Output format ('xlsx' or 'csv')
    
    Returns:
        str: Path to the created output file
    """
    # Load JSON data
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Extract metadata
    ausschreibung = data['metadata']['ausschreibung']
    bieter = data['metadata']['bieter']
    
    # Retrieve all bidder documents to identify unmatched ones
    all_bidder_docs = set()
    try:
        hochgeladene_docs = list_bidder_docs_json(ausschreibung, bieter, include_metadata=True)
        if isinstance(hochgeladene_docs, dict) and 'documents' in hochgeladene_docs:
            for doc in hochgeladene_docs['documents']:
                if isinstance(doc, dict) and 'name' in doc:
                    all_bidder_docs.add(doc['name'])
    except Exception as e:
        print(f"Warning: Could not retrieve bidder documents: {e}")
    
    # Prepare table data
    table_data = []
    gefordertes_doc_counter = 1
    
    # Collect all matched filenames to identify unmatched documents
    matched_filenames = set()
    
    for doc in data['matched_documents']:
        gefordertes_doc = doc['gefordertes_doc']
        
        # Get document info
        bezeichnung = gefordertes_doc['bezeichnung']
        beilage_nummer = gefordertes_doc.get('beilage_nummer', '')
        beschreibung = gefordertes_doc['beschreibung']
        
        # Add main row for "Gefordertes Dokument"
        table_data.append({
            'Typ': 'Gefordertes Dokument',
            'Nr': gefordertes_doc_counter,
            'Bezeichnung': bezeichnung,
            'Beilage': beilage_nummer if beilage_nummer else '',
            'Beschreibung': beschreibung.replace(',', ';'),
            'Dateiname': '',
            'Begründung': ''
        })
        
        # Handle matches - add as "Bieterdokument" rows
        if gefordertes_doc['matches']:
            for match in gefordertes_doc['matches']:
                matched_filenames.add(match['Dateiname'])
                table_data.append({
                    'Typ': 'Bieterdokument',
                    'Nr': '',
                    'Bezeichnung': '',
                    'Beilage': '',
                    'Beschreibung': '',
                    'Dateiname': match['Dateiname'],
                    'Begründung': match['Begründung'].replace(',', ';')
                })
        else:
            # No matches found
            table_data.append({
                'Typ': 'Bieterdokument',
                'Nr': '',
                'Bezeichnung': '',
                'Beilage': '',
                'Beschreibung': '',
                'Dateiname': 'NICHT GEFUNDEN',
                'Begründung': 'Kein passendes Dokument gefunden'
            })
        
        gefordertes_doc_counter += 1
    
    # Add unmatched documents section
    unmatched_docs = all_bidder_docs - matched_filenames
    if unmatched_docs:
        table_data.append({
            'Typ': 'Gefordertes Dokument',
            'Nr': gefordertes_doc_counter,
            'Bezeichnung': 'Unzuordenbare Dokumente',
            'Beilage': '',
            'Beschreibung': 'Dokumente die keinem geforderten Dokument zugeordnet werden konnten'.replace(',', ';'),
            'Dateiname': '',
            'Begründung': ''
        })

        for unmatched_doc in sorted(unmatched_docs):
            table_data.append({
                'Typ': 'Bieterdokument',
                'Nr': '',
                'Bezeichnung': '',
                'Beilage': '',
                'Beschreibung': '',
                'Dateiname': unmatched_doc,
                'Begründung': 'Nicht zugeordnet'
            })
    elif all_bidder_docs:
        # No unmatched docs, but we have the list
        table_data.append({
            'Typ': 'Gefordertes Dokument',
            'Nr': gefordertes_doc_counter,
            'Bezeichnung': 'Unzuordenbare Dokumente',
            'Beilage': '',
            'Beschreibung': 'Alle Bieterdokumente wurden zugeordnet'.replace(',', ';'),
            'Dateiname': '',
            'Begründung': ''
        })
    else:
        # Could not retrieve bidder documents
        table_data.append({
            'Typ': 'Gefordertes Dokument',
            'Nr': gefordertes_doc_counter,
            'Bezeichnung': 'Unzuordenbare Dokumente',
            'Beilage': '',
            'Beschreibung': 'Hinweis: Vollständige Liste aller Bieterdokumente nicht verfügbar'.replace(',', ';'),
            'Dateiname': '',
            'Begründung': ''
        })
    
    # Create DataFrame
    df = pd.DataFrame(table_data)
    
    # Generate output filename
    input_path = Path(json_file_path)
    output_filename = f"{ausschreibung}.{bieter}.{output_format}"
    output_path = input_path.parent / output_filename
    
    # Save to file
    if output_format.lower() == 'xlsx':
        # Create Excel file with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write header information
            header_df = pd.DataFrame({
                'Ausschreibung': [ausschreibung],
                'Bieter': [bieter]
            })
            header_df.to_excel(writer, sheet_name='Dokumentenmatching', 
                             index=False, startrow=0)
            
            # Write main data
            df.to_excel(writer, sheet_name='Dokumentenmatching', 
                       index=False, startrow=3)
            
            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Dokumentenmatching']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Apply formatting for better readability
            from openpyxl.styles import Font, PatternFill
            
            # Format "Gefordertes Dokument" rows with bold font and light background
            for row_num, row in enumerate(worksheet.iter_rows(min_row=4), start=4):
                if row[0].value == 'Gefordertes Dokument':
                    for cell in row:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
    
    elif output_format.lower() == 'csv':
        # Add header information to CSV
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(f"Ausschreibung: {ausschreibung}\n")
            f.write(f"Bieter: {bieter}\n")
            f.write("\n")

        # Append DataFrame to CSV with minimal quoting
        df.to_csv(output_path, mode='a', index=False, encoding='utf-8', quoting=csv.QUOTE_MINIMAL)
    
    return str(output_path)

def main():
    """
    Main function to handle command line arguments and execute conversion.
    """
    parser = argparse.ArgumentParser(
        description='Convert JSON document matching results to Excel/CSV format'
    )
    parser.add_argument('json_file', help='Path to the JSON file to convert')
    parser.add_argument('--format', '-f', choices=['xlsx', 'csv'], 
                       default='xlsx', help='Output format (default: xlsx)')
    
    args = parser.parse_args()
    
    # Check if input file exists
    if not Path(args.json_file).exists():
        print(f"Error: File '{args.json_file}' not found.")
        sys.exit(1)
    
    try:
        output_file = convert_json_to_table(args.json_file, args.format)
        print(f"Successfully converted to: {output_file}")
    except Exception as e:
        print(f"Error during conversion: {str(e)}")
        sys.exit(1)

if __name__ == '__main__':
    main()